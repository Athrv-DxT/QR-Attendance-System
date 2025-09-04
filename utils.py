import openpyxl
import qrcode
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from database import Participant, db
import uuid

def process_excel_file(file_path):
    """Process uploaded Excel file and extract name and email"""
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())
            else:
                headers.append("")
        
        # Look for name and email columns
        name_col_idx = None
        email_col_idx = None
        
        for i, header in enumerate(headers):
            if 'name' in header:
                name_col_idx = i + 1  # openpyxl uses 1-based indexing
            elif 'email' in header or 'mail' in header:
                email_col_idx = i + 1
        
        if not name_col_idx or not email_col_idx:
            raise ValueError("Excel file must contain 'name' and 'email' columns")
        
        participants = []
        for row in sheet.iter_rows(min_row=2):  # Skip header row
            name_cell = row[name_col_idx - 1]  # Convert to 0-based
            email_cell = row[email_col_idx - 1]
            
            if name_cell.value and email_cell.value:
                participants.append({
                    'name': str(name_cell.value).strip(),
                    'email': str(email_cell.value).strip().lower()
                })
        
        return participants
    except Exception as e:
        raise Exception(f"Error processing Excel file: {str(e)}")

def generate_qr_code(participant_id, name):
    """Generate QR code for participant"""
    try:
        # Create QR data with participant ID
        qr_data = f"PARTICIPANT_ID:{participant_id}"
        
        # Generate QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        # Create QR code image
        qr_img = qr.make_image(fill_color="black", back_color="white")
        
        # Ensure directory exists
        qr_dir = os.path.join('static', 'qr_codes')
        os.makedirs(qr_dir, exist_ok=True)
        
        # Save QR code
        filename = f"qr_{participant_id}_{uuid.uuid4().hex[:8]}.png"
        qr_path = os.path.join(qr_dir, filename)
        qr_img.save(qr_path)
        
        # Return path with forward slashes for web compatibility
        return qr_path.replace('\\', '/')
    except Exception as e:
        raise Exception(f"Error generating QR code: {str(e)}")

def send_qr_email(participant, message, smtp_server, smtp_port, email_user, email_password):
    """Send QR code to participant via email"""
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = email_user
        msg['To'] = participant.email
        msg['Subject'] = "Your QR Code for Event Attendance"
        
        # Email body
        body = f"""
        Dear {participant.name},
        
        {message}
        
        Please find your QR code attached. Show this QR code at the event for attendance marking.
        
        Best regards,
        Event Team
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Attach QR code
        if os.path.exists(participant.qr_code_path):
            with open(participant.qr_code_path, 'rb') as f:
                qr_image = MIMEImage(f.read())
                qr_image.add_header('Content-Disposition', 'attachment', filename='qr_code.png')
                msg.attach(qr_image)
        
        # Send email
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(email_user, email_password)
        server.send_message(msg)
        server.quit()
        
        return True
    except Exception as e:
        print(f"Error sending email to {participant.email}: {str(e)}")
        return False

def create_attendance_excel(attendances):
    """Create Excel file with attendance data"""
    try:
        data = []
        for attendance in attendances:
            data.append({
                'Name': attendance.participant.name,
                'Email': attendance.participant.email,
                'Entry Time': attendance.entry_time.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Attendance Report"
        
        # Add headers
        headers = ['Name', 'Email', 'Entry Time']
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Add data
        for row, record in enumerate(data, 2):
            worksheet.cell(row=row, column=1, value=record['Name'])
            worksheet.cell(row=row, column=2, value=record['Email'])
            worksheet.cell(row=row, column=3, value=record['Entry Time'])
        
        # Ensure uploads directory exists
        uploads_dir = 'uploads'
        os.makedirs(uploads_dir, exist_ok=True)
        
        filename = f"attendance_{uuid.uuid4().hex[:8]}.xlsx"
        filepath = os.path.join(uploads_dir, filename)
        workbook.save(filepath)
        
        return filepath
    except Exception as e:
        raise Exception(f"Error creating attendance Excel: {str(e)}")